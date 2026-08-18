"""Выгрузка отчёта: CSV (один файл, блоки подряд) и Excel (.xlsx, лист на блок).

Таблицы собираются один раз (``report_tables``) и пишутся двумя способами:
CSV — как выгрузка журнала (BOM, «;», массы целыми килограммами, доли с
запятой — открывается в русском Excel двойным щелчком); .xlsx — настоящая
книга через openpyxl с числовыми форматами (килограммы, тонны, проценты),
жирной шапкой и закреплённой первой строкой.
"""

from __future__ import annotations

import csv
import io
from dataclasses import dataclass, field
from datetime import UTC, datetime
from typing import Any
from zoneinfo import ZoneInfo

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from center import reports
from center.web import report_view

BISHKEK = ZoneInfo("Asia/Bishkek")

# форматы ячеек Excel по видам колонок
FMT_INT = "#,##0"
FMT_T = "#,##0.0"
FMT_PCT = "0.0%"

ColumnKind = str  # "text" | "int" | "kg" | "t" | "pct" | "dt"


@dataclass
class Table:
    """Один блок отчёта: заголовок, шапка, строки и виды колонок (для Excel)."""

    title: str
    headers: list[str]
    kinds: list[ColumnKind]
    rows: list[list[Any]] = field(default_factory=list)
    note: str = ""


def _dt(value: datetime | None) -> datetime | None:
    if value is None:
        return None
    if value.tzinfo is None:
        value = value.replace(tzinfo=UTC)
    return value.astimezone(BISHKEK).replace(tzinfo=None)


def _delta_text(delta: reports.Delta, *, tonnes: bool = False) -> str:
    """«10 → 12 (+20,0 %)»: начинается с числа — Excel не примет за формулу."""
    if tonnes:
        prev = report_view.fmt_tonnes(delta.previous)
        cur = report_view.fmt_tonnes(delta.current)
    else:
        prev = f"{delta.previous:,.0f}".replace(",", " ")
        cur = f"{delta.current:,.0f}".replace(",", " ")
    return f"{prev} → {cur} ({report_view.fmt_pct_change(delta)})"


def report_tables(report: reports.Report, *, site_name: str | None) -> list[Table]:
    """Все блоки отчёта в виде таблиц (общий источник для CSV и Excel)."""
    t = report.totals
    masses = report.masses
    cmp_ = report.comparison
    step = report.dynamics.step

    summary = Table(
        title="Итоги",
        headers=["Показатель", "Значение"],
        kinds=["text", "text"],
        rows=[
            ["Период", report.period.label],
            ["Объект", site_name or "все объекты"],
            ["Сформирован", report_view.generated_stamp(report.generated_at)],
            ["Взвешиваний", t.weighings],
            ["Тарирований", t.tarings],
            ["Офлайн-операций", t.offline],
            ["Доля офлайн-операций", report_view.fmt_share(t.offline_share)],
            ["Отказов команд АИС", t.refusals],
            ["Офлайн-операций без номера АИС", t.unlinked],
            ["Взвешиваний с тарой (есть нетто)", t.with_tare],
            ["Брутто по ним, кг", round(t.gross_with_tare_kg)],
            ["Тара по ним, кг", round(t.tare_sum_kg)],
            ["Нетто, кг", round(t.netto_kg)],
            ["Взвешиваний без тары (нет нетто)", t.without_tare],
            ["Брутто по ним, кг", round(t.gross_without_tare_kg)],
            ["Итого брутто, кг", round(t.gross_kg)],
            [
                "Средний брутто, кг",
                round(t.avg_gross_kg) if t.avg_gross_kg is not None else "—",
            ],
            *[
                [f"Без нетто: {reports.REASON_LABELS[reason]}", count]
                for reason, count in masses.reasons.items()
            ],
            ["Прошлый период", cmp_.period.label],
            ["Взвешиваний: прошлый период → этот", _delta_text(cmp_.weighings)],
            ["Тарирований: прошлый период → этот", _delta_text(cmp_.tarings)],
            ["Офлайн-операций: прошлый период → этот", _delta_text(cmp_.offline)],
            ["Нетто, т: прошлый период → этот", _delta_text(cmp_.netto_kg, tonnes=True)],
        ],
    )

    by_site = Table(
        title="По объектам",
        headers=[
            "Объект",
            "Взвешиваний",
            "Тарирований",
            "Офлайн",
            "Доля офлайн",
            "Нетто, т",
            "Средний брутто, кг",
            "Доступность",
            "Инцидентов",
            "Отказов АИС",
        ],
        kinds=["text", "int", "int", "int", "pct", "t", "kg", "pct", "int", "int"],
        rows=[
            [
                row.site_name,
                row.totals.weighings,
                row.totals.tarings,
                row.totals.offline,
                row.totals.offline_share,
                round(row.totals.netto_kg / 1000, 1),
                round(row.totals.avg_gross_kg) if row.totals.avg_gross_kg is not None else None,
                row.availability,
                row.incidents,
                row.totals.refusals,
            ]
            for row in report.sites
        ],
    )

    dynamics = Table(
        title=f"Динамика ({report_view.STEP_LABELS[step]})",
        headers=["Отрезок", "Взвешиваний", "Тарирований", "Офлайн", "Нетто, т"],
        kinds=["text", "int", "int", "int", "t"],
        rows=[
            [
                report_view.bucket_title(point.bucket, step),
                point.totals.weighings,
                point.totals.tarings,
                point.totals.offline,
                round(point.totals.netto_kg / 1000, 1),
            ]
            for point in report.dynamics.points
        ],
    )

    manual = Table(
        title="Ручные (офлайн) операции",
        headers=[
            "Объект",
            "Оператор",
            "Взвешиваний",
            "Тарирований",
            "Всего",
            "Доля от операций объекта",
            "Без номера АИС",
            f"Без номера старше {reports.UNLINKED_STALE_DAYS} сут",
        ],
        kinds=["text", "text", "int", "int", "int", "pct", "int", "int"],
        rows=[
            [
                row.site_name,
                row.operator or "(не указан)",
                row.weighings,
                row.tarings,
                row.operations,
                row.share,
                row.unlinked,
                row.unlinked_stale,
            ]
            for row in report.manual
        ],
    )

    reliability = Table(
        title="Надёжность",
        headers=[
            "Объект",
            "Доступность",
            "Уходов в офлайн",
            "Офлайн, мин",
            "Инцидентов индикатора",
            "Инцидентов камер",
            "Прочих инцидентов",
            "Отказов АИС",
            "Коды отказов",
        ],
        kinds=["text", "pct", "int", "int", "int", "int", "int", "int", "text"],
        rows=[
            [
                row.site_name,
                row.availability,
                row.offline_count,
                round(row.offline_seconds / 60),
                row.indicator_incidents,
                row.camera_incidents,
                row.other_incidents,
                row.refusals_total,
                ", ".join(f"{code} × {n}" for code, n in row.refusals.items()),
            ]
            for row in report.reliability
        ],
        note=(
            "Доступность — по событиям мониторинга (с 13.08.2026), "
            "для весов без агента не считается."
        ),
    )

    retare = Table(
        title="Сцепки к перетарированию",
        headers=[
            "Номер ТС",
            "Прицеп",
            "Взвешиваний без нетто",
            "Последнее взвешивание",
            "Последнее тарирование",
            "Тара, кг",
        ],
        kinds=["text", "text", "int", "dt", "dt", "kg"],
        rows=[
            [
                row.vehicle_number,
                row.trailer_number or "",
                row.weighings,
                _dt(row.last_weighed_at),
                _dt(row.last_tared_at),
                round(row.last_tare_value) if row.last_tare_value is not None else None,
            ]
            for row in masses.retare
        ],
        note=(
            f"Показаны первые {len(masses.retare)} из {masses.retare_total}"
            if masses.retare_total > len(masses.retare)
            else ""
        ),
    )
    return [summary, by_site, dynamics, manual, reliability, retare]


# --- CSV ---------------------------------------------------------------------


def _csv_cell(value: Any, kind: ColumnKind) -> str:
    if value is None:
        return ""
    if kind == "pct":
        return report_view.fmt_share(float(value)).replace(" %", "")
    if kind == "t":
        return f"{float(value):.1f}".replace(".", ",")
    if kind == "dt":
        return value.strftime("%d.%m.%Y %H:%M") if isinstance(value, datetime) else str(value)
    if isinstance(value, float):
        return f"{value:.0f}"
    text = str(value).strip()
    if text[:1] in {"=", "+", "-", "@"}:  # защита от формул Excel (как в журнале)
        return "'" + text
    return text


def report_csv(tables: list[Table]) -> bytes:
    """Все блоки в одном CSV: заголовок блока, шапка, строки, пустая строка."""
    buffer = io.StringIO()
    writer = csv.writer(buffer, delimiter=";", lineterminator="\r\n")
    for table in tables:
        writer.writerow([table.title])
        writer.writerow(table.headers)
        for row in table.rows:
            writer.writerow([_csv_cell(v, k) for v, k in zip(row, table.kinds, strict=True)])
        if table.note:
            writer.writerow([table.note])
        writer.writerow([])
    return ("\ufeff" + buffer.getvalue()).encode("utf-8")  # BOM: Excel понимает UTF-8


# --- Excel -------------------------------------------------------------------

_HEADER_FILL = PatternFill("solid", fgColor="EAF3EF")
_HEADER_FONT = Font(bold=True, color="0B5E43")
_SHEET_TITLE_BAD = set("[]:*?/\\")


def _sheet_name(title: str) -> str:
    name = "".join(ch for ch in title if ch not in _SHEET_TITLE_BAD)
    return name[:31] or "Лист"


def report_xlsx(tables: list[Table], *, title: str) -> bytes:
    """Книга Excel: лист на блок отчёта, форматы чисел по видам колонок."""
    book = Workbook()
    default = book.active
    assert default is not None
    book.remove(default)
    for table in tables:
        sheet = book.create_sheet(_sheet_name(table.title))
        sheet.append([title])
        sheet.cell(row=1, column=1).font = Font(bold=True, size=12)
        sheet.append([table.title])
        sheet.cell(row=2, column=1).font = Font(bold=True)
        sheet.append(table.headers)
        for col in range(1, len(table.headers) + 1):
            cell = sheet.cell(row=3, column=col)
            cell.fill = _HEADER_FILL
            cell.font = _HEADER_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        first_data = 4
        text_columns = [i for i, kind in enumerate(table.kinds, start=1) if kind == "text"]
        for row_index, row in enumerate(table.rows, start=first_data):
            sheet.append(list(row))
            # текст из данных (номера ТС, операторы, названия) — всегда строка:
            # значение вроде «=HYPERLINK(...)» openpyxl иначе записал бы формулой
            for col in text_columns:
                cell = sheet.cell(row=row_index, column=col)
                if isinstance(cell.value, str) and cell.value[:1] in {"=", "+", "-", "@"}:
                    cell.data_type = "s"
        for col, kind in enumerate(table.kinds, start=1):
            letter = get_column_letter(col)
            width = max(
                [len(str(table.headers[col - 1]))]
                + [len(str(r[col - 1])) for r in table.rows if r[col - 1] is not None]
            )
            sheet.column_dimensions[letter].width = min(max(width + 2, 10), 48)
            fmt = None
            if kind in ("int", "kg"):
                fmt = FMT_INT
            elif kind == "t":
                fmt = FMT_T
            elif kind == "pct":
                fmt = FMT_PCT
            elif kind == "dt":
                fmt = "DD.MM.YYYY HH:MM"
            if fmt is None:
                continue
            for row_index in range(first_data, first_data + len(table.rows)):
                sheet.cell(row=row_index, column=col).number_format = fmt
        if table.note:
            sheet.append([])
            sheet.append([table.note])
        sheet.freeze_panes = "A4"
    stream = io.BytesIO()
    book.save(stream)
    return stream.getvalue()
