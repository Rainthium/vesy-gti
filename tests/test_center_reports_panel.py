"""Тесты экрана «Отчёты» панели (маршруты /panel/reports*, этап 4) и его
выгрузок, разбора параметров и SVG-графиков.

Стенд — посев test_center_reports (объекты Альфа/Бета/Гамма) поверх
приложения панели как в test_center_panel. Проверяется:
- вход обязателен (303 на вход), страница со всеми блоками и числами посева;
- пресеты и произвольные даты, кривые параметры → период по умолчанию,
  будущее обрезается, ссылка «по объектам» переключает разбивку;
- ограничение диспетчера объектом: числа только своего объекта, селектора
  объектов нет, чужой site_id в адресе не расширяет видимость;
- печатная версия; CSV (BOM, «;», блоки) и Excel (листы, числа, форматы);
- report_view.resolve_query и charts — чистые функции.
"""

from __future__ import annotations

import io
from collections.abc import Iterator
from dataclasses import dataclass
from datetime import date
from pathlib import Path

import pytest
from fastapi import FastAPI
from fastapi.testclient import TestClient
from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.orm import Session, sessionmaker
from starlette.middleware.sessions import SessionMiddleware

from center import reports
from center.agents_ws.hub import AgentHub
from center.db.models import User
from center.web import charts, report_view
from center.web.router import create_panel_router
from tests.test_center_panel import (
    PANEL_LOGIN,
    PANEL_PASSWORD,
    _add_user,
    db,  # noqa: F401 — фикстуры БД панели
    panel_db_engine,  # noqa: F401
    panel_db_url,  # noqa: F401
)
from tests.test_center_reports import (
    Seed,
    seed,  # noqa: F401 — посев отчётов
)


@dataclass
class ReportsEnv:
    client: TestClient
    factory: sessionmaker[Session]
    seed: Seed


@pytest.fixture
def env(db: sessionmaker[Session], seed: Seed, tmp_path: Path) -> Iterator[ReportsEnv]:  # noqa: F811
    with db() as session:
        _add_user(session)
    app = FastAPI()
    app.add_middleware(SessionMiddleware, secret_key="test-secret", session_cookie="ves_test")
    app.include_router(create_panel_router(db, AgentHub(), photos_dir=tmp_path / "photos"))
    client = TestClient(app)
    yield ReportsEnv(client, db, seed)
    client.close()


def _login(env: ReportsEnv) -> None:
    response = env.client.post(
        "/panel/login",
        data={"login": PANEL_LOGIN, "password": PANEL_PASSWORD},
        follow_redirects=False,
    )
    assert response.status_code == 303


def _bind_to_site(env: ReportsEnv, site_id: int) -> None:
    with env.factory() as session:
        user = session.execute(select(User).where(User.login == PANEL_LOGIN)).scalar_one()
        user.site_id = site_id
        session.commit()


PERIOD_QS = "date_from=2026-08-01&date_to=2026-08-10"


class TestAccess:
    @pytest.mark.parametrize(
        "path",
        [
            "/panel/reports",
            "/panel/reports/print",
            "/panel/reports/export.csv",
            "/panel/reports/export.xlsx",
        ],
    )
    def test_requires_login(self, env: ReportsEnv, path: str) -> None:
        response = env.client.get(path, follow_redirects=False)
        assert response.status_code == 303
        assert response.headers["location"].startswith("/panel/login")

    def test_header_tab_present(self, env: ReportsEnv) -> None:
        _login(env)
        page = env.client.get("/panel/journal").text
        assert 'href="/panel/reports"' in page


class TestReportPage:
    def test_page_shows_seed_numbers(self, env: ReportsEnv) -> None:
        _login(env)
        response = env.client.get(f"/panel/reports?{PERIOD_QS}")
        assert response.status_code == 200
        page = response.text
        assert "01.08.2026 — 10.08.2026" in page
        # блоки экрана
        for title in (
            "Массы за период",
            "По объектам",
            "Динамика",
            "Ручные (офлайн) операции",
            "Надёжность",
            "Сцепки к перетарированию",
        ):
            assert title in page, title
        # числа посева: 9 взвешиваний, нетто 69,0 т, офлайн 1, средний брутто
        assert 'class="kpi-v">9<' in page
        assert "69,0" in page
        assert "Оператор А" in page
        assert "СВХ «Альфа»" in page and "ПЗТК «Бета»" in page and "СВХ «Гамма»" in page
        # причины без нетто и сцепки
        assert "тарирование устарело: 1" in page
        assert "действующего тарирования не было: 2" in page
        assert "V3" in page and "V2" in page
        # отказы АИС по кодам
        assert "ERR_UNSTABLE" in page and "ERR_VEHICLE_TIMEOUT" in page
        # графики — inline SVG
        assert page.count("<svg") == 4
        # ссылки печати/экспорта несут те же фильтры
        assert "/panel/reports/print?date_from=2026-08-01&amp;date_to=2026-08-10" in page
        assert "/panel/reports/export.xlsx?date_from=2026-08-01&amp;date_to=2026-08-10" in page

    def test_default_and_presets(self, env: ReportsEnv) -> None:
        _login(env)
        page = env.client.get("/panel/reports").text
        # по умолчанию — этот месяц: чип активен, скрытое поле пресета заполнено
        assert 'class="chip active"' in page and 'name="preset" value="month"' in page
        page = env.client.get("/panel/reports?preset=year").text
        assert "по месяцам" in page  # шаг динамики по длине периода
        # мусорный пресет и даты → период по умолчанию, не 500
        response = env.client.get("/panel/reports?preset=nope&date_from=abc&date_to=")
        assert response.status_code == 200 and 'value="month"' in response.text

    def test_custom_dates_swap_and_future_clamp(self, env: ReportsEnv) -> None:
        _login(env)
        page = env.client.get("/panel/reports?date_from=2026-08-10&date_to=2026-08-01").text
        assert "01.08.2026 — 10.08.2026" in page  # порядок дат исправлен
        response = env.client.get("/panel/reports?date_from=2099-01-01&date_to=2099-12-31")
        assert response.status_code == 200
        today = date.today().strftime("%d.%m.%Y")  # будущее обрезано до сегодня
        assert today in response.text and "2099" not in response.text.split("<main")[1]

    def test_split_toggle(self, env: ReportsEnv) -> None:
        _login(env)
        page = env.client.get(f"/panel/reports?{PERIOD_QS}").text
        assert "split=sites" in page  # ссылка «По объектам»
        split = env.client.get(f"/panel/reports?{PERIOD_QS}&split=sites").text
        # линии по объектам: легенда с именами внутри SVG
        assert split.count("<svg") == 4
        assert "Суммарно" in split
        assert "СВХ «Альфа»" in split.split("Динамика")[1]

    def test_single_site_filter(self, env: ReportsEnv) -> None:
        _login(env)
        page = env.client.get(f"/panel/reports?{PERIOD_QS}&site_id={env.seed.site_b}").text
        assert "ПЗТК «Бета»" in page
        assert 'class="kpi-v">2<' in page  # два взвешивания Беты
        # у одного объекта переключателя «по объектам» нет
        assert "split=sites" not in page.split("Динамика")[1]
        # чужой/несуществующий объект → как «все объекты»
        page = env.client.get(f"/panel/reports?{PERIOD_QS}&site_id=99999").text
        assert 'class="kpi-v">9<' in page


class TestScope:
    def test_dispatcher_sees_only_own_site(self, env: ReportsEnv) -> None:
        _login(env)
        _bind_to_site(env, env.seed.site_a)
        page = env.client.get(f"/panel/reports?{PERIOD_QS}").text
        assert 'class="kpi-v">7<' in page  # только Альфа
        assert "ПЗТК «Бета»" not in page
        assert '<select name="site_id">' not in page
        assert "СВХ «Альфа»" in page

    def test_scope_cannot_be_widened_by_query(self, env: ReportsEnv) -> None:
        _login(env)
        _bind_to_site(env, env.seed.site_a)
        page = env.client.get(f"/panel/reports?{PERIOD_QS}&site_id={env.seed.site_b}").text
        assert 'class="kpi-v">7<' in page
        assert "ПЗТК «Бета»" not in page
        csv_body = env.client.get(
            f"/panel/reports/export.csv?{PERIOD_QS}&site_id={env.seed.site_b}"
        )
        assert "Бета" not in csv_body.text
        assert "Альфа" in csv_body.text


class TestPrintAndExports:
    def test_print_page(self, env: ReportsEnv) -> None:
        _login(env)
        response = env.client.get(f"/panel/reports/print?{PERIOD_QS}")
        assert response.status_code == 200
        page = response.text
        assert "Отчёт по взвешиваниям за 01.08.2026 — 10.08.2026 — все объекты" in page
        assert 'class="panel print-page"' in page
        assert "window.print()" in page
        assert 'href="/panel/journal"' not in page  # без шапки панели

    def test_csv_export(self, env: ReportsEnv) -> None:
        _login(env)
        response = env.client.get(f"/panel/reports/export.csv?{PERIOD_QS}")
        assert response.status_code == 200
        assert response.headers["content-type"].startswith("text/csv")
        assert (
            'filename="otchet_2026-08-01_2026-08-10.csv"' in response.headers["content-disposition"]
        )
        body = response.content.decode("utf-8")
        assert body.startswith("\ufeff")
        text = body.lstrip("\ufeff")
        assert text.startswith("Итоги\r\nПоказатель;Значение\r\n")
        for block in (
            "По объектам",
            "Динамика (по дням)",
            "Ручные (офлайн) операции",
            "Надёжность",
            "Сцепки к перетарированию",
        ):
            assert f"\r\n{block}\r\n" in text, block
        assert "Взвешиваний;9\r\n" in text
        assert "Нетто, кг;69000\r\n" in text
        # строка объекта: доли с запятой, тонны с одним знаком
        assert "СВХ «Альфа»;7;0;1;14,3;27,0;" in text
        # оператор ручного режима и сцепка к перетарированию
        assert "Оператор А" in text and "V3;;1;" in text

    def test_xlsx_export(self, env: ReportsEnv) -> None:
        _login(env)
        response = env.client.get(f"/panel/reports/export.xlsx?{PERIOD_QS}")
        assert response.status_code == 200
        assert response.headers["content-type"].startswith(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        book = load_workbook(io.BytesIO(response.content))
        assert book.sheetnames == [
            "Итоги",
            "По объектам",
            "Динамика (по дням)",
            "Ручные (офлайн) операции",
            "Надёжность",
            "Сцепки к перетарированию",
        ]
        summary = book["Итоги"]
        assert (
            summary["A1"].value == "Отчёт по взвешиваниям за 01.08.2026 — 10.08.2026 — все объекты"
        )
        rows = {row[0].value: row[1].value for row in summary.iter_rows(min_row=4)}
        assert rows["Взвешиваний"] == 9 and rows["Нетто, кг"] == 69000
        assert rows["Без нетто: тарирование устарело"] == 1
        by_site = book["По объектам"]
        header = [c.value for c in by_site[3]]
        assert header[:3] == ["Объект", "Взвешиваний", "Тарирований"]
        alpha = next(
            r for r in by_site.iter_rows(min_row=4, values_only=True) if r[0] == "СВХ «Альфа»"
        )
        assert alpha[1] == 7 and alpha[4] == pytest.approx(1 / 7) and alpha[5] == 27.0
        # процентный формат у доли офлайн, закреплённая шапка
        assert by_site["E4"].number_format == "0.0%"
        assert by_site.freeze_panes == "A4"
        dynamics = book["Динамика (по дням)"]
        assert dynamics.max_row - 3 == 10  # десять отрезков

    def test_xlsx_text_is_never_formula(self, env: ReportsEnv) -> None:
        """Номер ТС «=HYPERLINK(...)» из ручного ввода не должен стать формулой Excel."""
        from datetime import UTC, datetime

        from tests.test_center_reports import _weighing

        with env.factory() as session:
            _weighing(
                session,
                env.seed.scale_a,
                datetime(2026, 8, 9, 6, 0, tzinfo=UTC),
                massa=12345,
                vehicle='=HYPERLINK("http://evil")',
            )
        _login(env)
        response = env.client.get(f"/panel/reports/export.xlsx?{PERIOD_QS}")
        book = load_workbook(io.BytesIO(response.content))
        sheet = book["Сцепки к перетарированию"]
        cells = [
            c
            for row in sheet.iter_rows(min_row=4)
            for c in row
            if str(c.value).upper()
            == '=HYPERLINK("HTTP://EVIL")'  # ключ сцепки в отчёте — верхний регистр
        ]
        assert cells, "строка сцепки не найдена"
        assert all(c.data_type == "s" for c in cells)
        # и в CSV — с апострофом, как в журнале
        csv_text = env.client.get(f"/panel/reports/export.csv?{PERIOD_QS}").text
        assert "'=HYPERLINK" in csv_text and ";=HYPERLINK" not in csv_text

    def test_garbage_dates_do_not_500(self, env: ReportsEnv) -> None:
        _login(env)
        for qs in (
            "date_from=0001-01-01",
            "date_from=0001-01-01&date_to=0001-01-31",
            "date_from=9999-12-31",
        ):
            response = env.client.get(f"/panel/reports?{qs}")
            assert response.status_code == 200, qs


class TestResolveQuery:
    def test_preset_wins_over_dates(self) -> None:
        q = report_view.resolve_query(
            preset="7d",
            date_from="2026-01-01",
            date_to="2026-01-31",
            site_id=None,
            split=None,
            today=date(2026, 8, 18),
        )
        assert q.preset == "7d" and q.period == reports.Period(date(2026, 8, 12), date(2026, 8, 18))
        assert q.unit_months is None
        assert q.query_string() == "preset=7d"

    def test_custom_dates(self) -> None:
        q = report_view.resolve_query(
            preset="",
            date_from="2026-08-10",
            date_to="2026-08-01",
            site_id=3,
            split="sites",
            today=date(2026, 8, 18),
        )
        assert q.preset == "custom" and q.period == reports.Period(
            date(2026, 8, 1), date(2026, 8, 10)
        )
        assert q.query_string() == "date_from=2026-08-01&date_to=2026-08-10&site_id=3&split=sites"
        assert (
            q.query_string(split="") == "date_from=2026-08-01&date_to=2026-08-10&site_id=3&split="
        )
        # одна дата — один день; будущее режется сегодняшним днём
        one = report_view.resolve_query(
            preset=None,
            date_from="2026-08-05",
            date_to=None,
            site_id=None,
            split=None,
            today=date(2026, 8, 18),
        )
        assert one.period == reports.Period(date(2026, 8, 5), date(2026, 8, 5))
        future = report_view.resolve_query(
            preset=None,
            date_from="2026-08-30",
            date_to="2026-09-10",
            site_id=None,
            split=None,
            today=date(2026, 8, 18),
        )
        assert future.period == reports.Period(date(2026, 8, 18), date(2026, 8, 18))
        # слишком длинный — обрезается с начала
        long = report_view.resolve_query(
            preset=None,
            date_from="2010-01-01",
            date_to="2026-08-18",
            site_id=None,
            split=None,
            today=date(2026, 8, 18),
        )
        assert long.period.days == report_view.MAX_PERIOD_DAYS

    def test_min_date_floor(self) -> None:
        early = report_view.resolve_query(
            preset=None,
            date_from="0001-01-01",
            date_to="0001-01-31",
            site_id=None,
            split=None,
            today=date(2026, 8, 18),
        )
        assert early.period == reports.Period(report_view.MIN_DATE, report_view.MIN_DATE)
        assert early.previous.date_from >= date(1999, 1, 1)  # арифметика периодов не падает
        one_sided = report_view.resolve_query(
            preset=None,
            date_from="0001-01-01",
            date_to=None,
            site_id=None,
            split=None,
            today=date(2026, 8, 18),
        )
        assert one_sided.period == reports.Period(report_view.MIN_DATE, report_view.MIN_DATE)

    def test_default_and_units(self) -> None:
        q = report_view.resolve_query(
            preset=None,
            date_from=None,
            date_to=None,
            site_id=None,
            split=None,
            today=date(2026, 8, 18),
        )
        assert q.preset == "month" and q.unit_months == 1
        assert q.previous == reports.Period(date(2026, 7, 1), date(2026, 7, 18))
        quarter = report_view.resolve_query(
            preset="quarter",
            date_from=None,
            date_to=None,
            site_id=None,
            split=None,
            today=date(2026, 8, 18),
        )
        assert quarter.unit_months == 3
        assert quarter.previous == reports.Period(date(2026, 4, 1), date(2026, 5, 18))

    def test_labels_and_formats(self) -> None:
        assert report_view.bucket_label(date(2026, 8, 3), "day") == "03.08"
        assert report_view.bucket_label(date(2026, 8, 1), "month") == "авг 2026"
        assert report_view.bucket_title(date(2026, 8, 3), "week") == "03.08.2026 — 09.08.2026"
        assert report_view.fmt_hours(0) == "—"
        assert report_view.fmt_hours(65 * 60) == "1 ч 05 мин"
        assert report_view.fmt_hours(12 * 60) == "12 мин"
        assert report_view.fmt_share(0.125) == "12,5 %"
        assert report_view.fmt_share(None) == "—"
        assert report_view.fmt_tonnes(128640.4) == "128,6"
        assert report_view.fmt_tonnes(1234567) == "1 234,6"
        assert report_view.fmt_delta(reports.Delta(12, 10)) == "+2 (+20,0 %)"
        assert report_view.fmt_delta(reports.Delta(8, 10)) == "−2 (−20,0 %)"
        assert report_view.fmt_delta(reports.Delta(5, 0)) == "+5 (прошлый период: 0)"
        assert report_view.fmt_delta(reports.Delta(69000, 60000), tonnes=True) == "+9,0 (+15,0 %)"
        assert report_view.fmt_pct_change(reports.Delta(0, 0)) == "—"


class TestCharts:
    def test_number_and_ceiling(self) -> None:
        assert charts.fmt_number(128640) == "128 640"
        assert charts.fmt_number(2.5) == "2,5"
        assert charts.fmt_number(2.0) == "2"
        assert charts.nice_ceiling(0) == 4
        assert charts.nice_ceiling(12) == 16
        assert charts.nice_ceiling(130) == 160
        assert charts.nice_ceiling(999) == 1000
        for value in (1, 7, 33, 999, 12345, 1e6):
            top = charts.nice_ceiling(value)
            assert top >= value and (top / 4) == pytest.approx(round(top / 4, 6))

    def test_bar_chart_escapes_and_scales(self) -> None:
        svg = charts.bar_chart_horizontal(
            [("СВХ <А>", 10.0), ("Б & В", 5.0), ("Пусто", 0.0)], unit="т", decimals=1, title="t"
        )
        assert svg.startswith("<svg") and svg.endswith("</svg>")
        assert "&lt;А&gt;" in svg and "&amp;" in svg and "<А>" not in svg
        assert "10,0 т" in svg and "0,0 т" in svg
        # столбик пустого значения нулевой ширины, максимальный — во всю шкалу
        assert 'width="0.0"' in svg and 'width="440.0"' in svg
        assert "нет данных" in charts.bar_chart_horizontal([])

    def test_column_and_line_charts(self) -> None:
        labels = [f"{d:02d}.08" for d in range(1, 41)]
        values = [float(i % 7) for i in range(40)]
        svg = charts.column_chart(labels, values, title="c")
        assert svg.count("<rect") == 40
        # подписи оси X прорежены (не больше 16), деления оси Y без дробей
        assert svg.count(">01.08<") == 1 and svg.count("<text") < 40 + 30
        assert ">8<" in svg  # верх шкалы 8 при максимуме 6 → 4 деления по 2
        line = charts.line_chart(["a", "b", "c"], [("Первый", [1, 2, 3]), ("Второй", [0, 0, 5])])
        assert line.count("<path") == 2 and "Первый" in line and "Второй" in line
        assert "нет данных" in charts.line_chart(["a"], [])
        with pytest.raises(ValueError):
            charts.column_chart(["a", "b"], [1.0])
